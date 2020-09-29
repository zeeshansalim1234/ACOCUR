import pandas as pd
import datetime
from sklearn.feature_extraction.text import CountVectorizer
from scipy.spatial import distance
import os

REVIEW_FILE_PATH = "../Output/Review/CTypes"
COMMIT_FILE_PATH = "../Output/Commit/CTypes"
OUTPUT_PATH = "../Output/FINAL"

def CombinedSummary(duration = 30):
    import xlsxwriter

    for file in os.listdir(REVIEW_FILE_PATH):

        review_filename = REVIEW_FILE_PATH + "/" + file
        commit_filename = COMMIT_FILE_PATH + "/" + file
        outputfile = OUTPUT_PATH + "/" + file.replace(".csv",".xlsx")

        workbook = xlsxwriter.Workbook(outputfile)
        worksheet = workbook.add_worksheet("Review Summary")
        worksheet.set_column(1, 1, 25)
        bold = workbook.add_format({'bold': True})

        df_Reviews = pd.read_csv(review_filename)
        total_count = len(df_Reviews)
        row = 1
        worksheet.write(row, 1, "Total number of reviews", bold)
        worksheet.write(row, 2, total_count)

        row += 2

        if total_count > 0:
            worksheet.write(row, 1, "Type", bold)
            worksheet.write(row, 2, "Count", bold)

            df_Type = df_Reviews['Type'].value_counts().rename_axis('Type').reset_index(name='Count')
            for i, r in df_Type.iterrows():
                row += 1
                worksheet.write(row, 1, r['Type'])
                worksheet.write(row, 2, r['Count'])

        worksheet1 = workbook.add_worksheet("Commit Summary")
        worksheet1.set_column(1, 1, 25)

        df_Commit = pd.read_csv(commit_filename)
        total_count = len(df_Commit)
        row = 1
        worksheet1.write(row, 1, "Total number of commits", bold)
        worksheet1.write(row, 2, total_count)

        row += 2

        if total_count > 0:
            worksheet1.write(row, 1, "Type", bold)
            worksheet1.write(row, 2, "Count", bold)
            df_Type = df_Commit['Type'].value_counts().rename_axis('Type').reset_index(name='Count')
            for i, r in df_Type.iterrows():
                row += 1
                worksheet1.write(row, 1, r['Type'])
                worksheet1.write(row, 2, r['Count'])

        workbook.close()

        from openpyxl import load_workbook

        book = load_workbook(outputfile)
        writer = pd.ExcelWriter(outputfile, engine='openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)


        df_Reviews['Date'] = pd.to_datetime(df_Reviews['Date'])
        df_Reviews['Date'] = df_Reviews['Date'].dt.date
        df_Commit['Date'] = (pd.to_datetime(df_Commit['Date'])).dt.date

        df_Reviews['#Linked_Commits'] = 0
        df_Reviews['Details'] = ""
        df_Reviews['Commit_Ref'] = ""

        for index, row in df_Reviews.iterrows():
            date_start = row['Date']
            date_end = date_start + datetime.timedelta(days=duration)

            df_related = df_Commit.loc[(df_Commit['Date'] >= date_start) & (df_Commit['Date'] <= date_end)]
            details = ""
            reference = ""
            if len(df_related) > 0:

                commit_ref = df_related['SL'].tolist()
                reference = ' ; '.join(map(str, commit_ref))

                df_Type = df_related['Type'].value_counts().rename_axis('Type').reset_index(name='Count')
                for i, r in df_Type.iterrows():
                    details += r['Type'] + " : " + str(r['Count']) + " ; "

            df_Reviews.at[index, '#Linked_Commits'] = len(df_related)
            df_Reviews.at[index, 'Details'] = details
            df_Reviews.at[index, 'Commit_Ref'] = reference

        df_Reviews.to_excel(writer, sheet_name='Reviews with Commits link', index=False,
                            header=['SL', 'Reviews', 'Date', 'Compatibility Type', '#Linked_Commits', 'Details', 'Commit_SL'])
        df_Commit.to_excel(writer, sheet_name='All Commits', index=False, header=['SL', 'Commit', 'Date', 'Compatibility Type'])
        writer.save()


def ReviewSummary():

    import xlsxwriter

    for file in os.listdir(REVIEW_FILE_PATH):
        review_filename = REVIEW_FILE_PATH + "/" + file
        outputfile = OUTPUT_PATH + "/" + file.replace(".csv", ".xlsx")

        workbook = xlsxwriter.Workbook(outputfile)
        worksheet = workbook.add_worksheet("Review Summary")
        worksheet.set_column(1, 1, 25)
        bold = workbook.add_format({'bold': True})

        df_Reviews = pd.read_csv(review_filename)
        total_count = len(df_Reviews)
        row = 1
        worksheet.write(row, 1, "Total number of reviews", bold)
        worksheet.write(row, 2, total_count)

        row += 2

        if total_count > 0:
            worksheet.write(row, 1, "Type", bold)
            worksheet.write(row, 2, "Count", bold)

            df_Type = df_Reviews['Type'].value_counts().rename_axis('Type').reset_index(name='Count')
            for i, r in df_Type.iterrows():
                row += 1
                worksheet.write(row, 1, r['Type'])
                worksheet.write(row, 2, r['Count'])

        workbook.close()

        Add_Sheet(df_Reviews, outputfile, 'Review Text', 'All Reviews')


def CommitSummary():

    import xlsxwriter

    for file in os.listdir(COMMIT_FILE_PATH):
        commit_filename = COMMIT_FILE_PATH + "/" + file
        outputfile = OUTPUT_PATH + "/" + file.replace(".csv", ".xlsx")

        workbook = xlsxwriter.Workbook(outputfile)
        worksheet = workbook.add_worksheet("Commit Summary")
        worksheet.set_column(1, 1, 25)
        bold = workbook.add_format({'bold': True})

        df_Commit = pd.read_csv(commit_filename)
        total_count = len(df_Commit)
        row = 1
        worksheet.write(row, 1, "Total number of commits", bold)
        worksheet.write(row, 2, total_count)

        row += 2

        if total_count > 0:
            worksheet.write(row, 1, "Type", bold)
            worksheet.write(row, 2, "Count", bold)
            df_Type = df_Commit['Type'].value_counts().rename_axis('Type').reset_index(name='Count')
            for i, r in df_Type.iterrows():
                row += 1
                worksheet.write(row, 1, r['Type'])
                worksheet.write(row, 2, r['Count'])

        workbook.close()

        Add_Sheet(df_Commit, outputfile, 'Commit Text', 'All Commits')


def Add_Sheet(df, file, headertext, sheetname):
    from openpyxl import load_workbook

    book = load_workbook(file)
    writer = pd.ExcelWriter(file, engine='openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)

    df.to_excel(writer, sheet_name=sheetname, index=False,
                        header=['SL', headertext, 'Date', 'Compatibility Type'])
    writer.save()